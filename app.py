#!/usr/bin/env python3
"""
CBSE School Scraper — Streamlit Dashboard
Interactive web interface to search and download CBSE affiliated school data.
"""

# import subprocess
# import sys
# from pathlib import Path

# # Run setup.sh automatically before the app starts
# _setup_script = Path(__file__).parent / "setup.sh"
# if _setup_script.exists():
#     subprocess.run(["bash", str(_setup_script)], check=True)

import io
import os
import re
import time
import urllib.parse
from collections import Counter

import streamlit as st
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import (
    NoSuchElementException,
    StaleElementReferenceException,
)
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SARAS_URL = "https://saras.cbse.gov.in/saras/AffiliatedList/ListOfSchdirReport"

# ──────────────────────── Page config ──────────────────────────────────────

st.set_page_config(
    page_title="CBSE School Scraper",
    page_icon="🏫",
    layout="wide",
)

# ──────────────────────── Custom CSS ───────────────────────────────────────

st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1F4E79, #2980b9);
        color: white;
        padding: 1.5rem 2rem;
        border-radius: 10px;
        margin-bottom: 1.5rem;
        text-align: center;
    }
    .main-header h1 { color: white; margin: 0; font-size: 2rem; }
    .main-header p { color: #d4e6f1; margin: 0.3rem 0 0; font-size: 1rem; }
    .stat-card {
        background: #f0f7ff;
        border-left: 4px solid #1F4E79;
        padding: 1rem;
        border-radius: 6px;
        margin-bottom: 1rem;
    }
    .stat-card h3 { margin: 0; color: #1F4E79; }
    .stat-card p { margin: 0; font-size: 1.8rem; font-weight: bold; color: #2c3e50; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class='main-header'>
    <h1>🏫 CBSE School Scraper</h1>
    <p>Search and download CBSE affiliated school data from SARAS 7.0</p>
</div>
""", unsafe_allow_html=True)


# ──────────────────── Fetch states/districts from SARAS ────────────────────

def _open_state_page(driver):
    """Navigate to the SARAS page and activate the State-wise radio.
    Returns the state dropdown element.
    """
    driver.get(SARAS_URL)
    time.sleep(5)
    radio = driver.find_element(By.ID, "SearchMainRadioState_wise")
    driver.execute_script("arguments[0].click()", radio)
    time.sleep(2)
    return driver.find_element(By.ID, "State")


def _collect_states(driver):
    """Return {state_name: state_id} from the currently visible dropdown."""
    sel = _open_state_page(driver)
    opts = sel.find_elements(By.TAG_NAME, "option")
    state_map = {}
    for o in opts:
        val = o.get_attribute("value")
        txt = o.text.strip()
        if val and txt and txt not in ("--select--", "--Select--", ""):
            state_map[txt] = val
    return state_map


def _fetch_districts_for_states(state_items):
    """Open a fresh browser session and fetch districts for a list of
    (state_name, state_id) pairs.  Returns {state_name: [district, ...]}."""
    driver = _create_headless_driver()
    try:
        sel = _open_state_page(driver)
        result = {}
        for state_name, state_id in state_items:
            try:
                sel = driver.find_element(By.ID, "State")
                Select(sel).select_by_value(state_id)
                driver.execute_script(
                    "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));",
                    sel,
                )
                time.sleep(1.5)

                dist_sel = driver.find_element(By.ID, "District")
                dist_opts = dist_sel.find_elements(By.TAG_NAME, "option")
                districts = []
                for d in dist_opts:
                    dt = d.text.strip()
                    if dt and dt not in ("--Select--", "--select--", ""):
                        districts.append(dt)
                result[state_name] = sorted(districts)
            except Exception:
                # If a single state fails, skip and let the retry batch handle it
                continue
        return result
    finally:
        driver.quit()


_BATCH_SIZE = 8  # states per browser session to avoid long-lived sessions


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_states_and_districts():
    """Fetch all states and their districts from the SARAS website.
    Uses short-lived browser sessions (one per batch of states) to avoid
    Chrome crashing from prolonged sessions.

    Returns (state_map, district_map):
        state_map:    { "STATE_NAME": "state_id", ... }
        district_map: { "STATE_NAME": ["DISTRICT1", "DISTRICT2", ...], ... }
    """
    # Step 1: Collect state list (quick — single page load)
    driver = _create_headless_driver()
    try:
        state_map = _collect_states(driver)
    finally:
        driver.quit()

    if not state_map:
        raise RuntimeError("Could not load any states from the SARAS website.")

    # Step 2: Fetch districts in small batches, restarting the browser each time
    items = list(state_map.items())
    district_map = {}

    for i in range(0, len(items), _BATCH_SIZE):
        batch = items[i : i + _BATCH_SIZE]
        for attempt in range(3):
            try:
                batch_result = _fetch_districts_for_states(batch)
                district_map.update(batch_result)
                break
            except Exception:
                if attempt == 2:
                    raise
                time.sleep(2)

    # Retry any states that got skipped
    missing = [
        (name, sid) for name, sid in items if name not in district_map
    ]
    if missing:
        for attempt in range(2):
            try:
                retry_result = _fetch_districts_for_states(missing)
                district_map.update(retry_result)
                break
            except Exception:
                time.sleep(2)

    return state_map, district_map


# ──────────────────── Selenium helpers ─────────────────────────────────────

# Known Chromium binary locations on Ubuntu / Lightning AI
_CHROME_CANDIDATES = [
    "/usr/bin/google-chrome-stable",
    "/usr/bin/google-chrome",
    "/usr/bin/chromium-browser",
    "/usr/bin/chromium",
    "/opt/google/chrome/google-chrome",
    "/usr/local/bin/google-chrome",
]
# Known chromedriver locations when installed via apt / setup.sh
_CHROMEDRIVER_CANDIDATES = [
    "/usr/bin/chromedriver",
    "/usr/local/bin/chromedriver",
    "/usr/lib/chromium-browser/chromedriver",
    "/usr/lib/chromium/chromedriver",
    "/snap/bin/chromium.chromedriver",
    "/opt/google/chrome/chromedriver",
]


def _create_headless_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-setuid-sandbox")
    options.add_argument("--disable-extensions")
    options.add_argument("--remote-debugging-port=0")
    # Memory / stability flags to prevent early crashes
    options.add_argument("--disable-background-networking")
    options.add_argument("--disable-default-apps")
    options.add_argument("--disable-sync")
    options.add_argument("--disable-translate")
    options.add_argument("--metrics-recording-only")
    options.add_argument("--no-first-run")
    options.add_argument("--safebrowsing-disable-auto-update")
    options.add_argument("--disable-background-timer-throttling")
    options.add_argument("--disable-renderer-backgrounding")
    options.add_argument("--disable-backgrounding-occluded-windows")
    options.page_load_strategy = "eager"

    # Locate Chrome/Chromium binary
    chrome_bin = os.environ.get("CHROME_BIN", "")
    if not chrome_bin:
        for candidate in _CHROME_CANDIDATES:
            if os.path.exists(candidate):
                chrome_bin = candidate
                break
    # Also try `which` in case binary is in a non-standard PATH location
    if not chrome_bin:
        import shutil
        for name in ("google-chrome-stable", "google-chrome", "chromium-browser", "chromium"):
            found = shutil.which(name)
            if found:
                chrome_bin = found
                break
    if chrome_bin:
        options.binary_location = chrome_bin

    # Prefer system-installed chromedriver (version-matched to system Chromium)
    # over webdriver-manager, which may download a mismatched version.
    chromedriver_path = os.environ.get("CHROMEDRIVER_PATH", "")
    if not chromedriver_path:
        for candidate in _CHROMEDRIVER_CANDIDATES:
            if os.path.exists(candidate):
                chromedriver_path = candidate
                break
    if not chromedriver_path:
        import shutil
        found = shutil.which("chromedriver")
        if found:
            chromedriver_path = found

    if chromedriver_path:
        service = Service(chromedriver_path)
    else:
        # Fallback: let webdriver-manager download a matching chromedriver.
        # This handles Streamlit Community Cloud, local Mac dev, and any
        # environment where system chromedriver isn't at a known path.
        service = Service(ChromeDriverManager().install())

    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(10)
    return driver


def _js_click(driver, element):
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
    time.sleep(0.3)
    driver.execute_script("arguments[0].click();", element)


def _select_option_by_text(driver, select_id, value):
    sel = driver.find_element(By.ID, select_id)
    opts = sel.find_elements(By.TAG_NAME, "option")
    for opt in opts:
        if opt.text.strip().upper() == value.upper():
            Select(sel).select_by_visible_text(opt.text.strip())
            driver.execute_script(
                "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));",
                sel,
            )
            return True
    for opt in opts:
        if value.upper() in opt.text.strip().upper():
            Select(sel).select_by_visible_text(opt.text.strip())
            driver.execute_script(
                "arguments[0].dispatchEvent(new Event('change', {bubbles: true}));",
                sel,
            )
            return True
    return False


# ──────────────────── Scraping logic ───────────────────────────────────────

def _get_total_entries(driver):
    try:
        info = driver.find_element(By.ID, "myTable_info")
        m = re.search(r"of\s+([\d,]+)\s+", info.text, re.I)
        if m:
            return int(m.group(1).replace(",", ""))
    except NoSuchElementException:
        pass
    return 0


def _parse_row(row):
    cells = row.find_elements(By.TAG_NAME, "td")
    if len(cells) < 6:
        return None

    def txt(idx):
        try:
            return cells[idx].text.strip()
        except (IndexError, StaleElementReferenceException):
            return ""

    s_no = txt(0)

    aff_text = txt(1)
    aff_no = school_code = ""
    m = re.search(r"Aff\.?\s*No\.?\s*:?\s*(\S+)", aff_text, re.I)
    if m:
        aff_no = m.group(1).strip().rstrip(",")
    m = re.search(r"Sch\.?\s*Code\s*:?\s*(\S+)", aff_text, re.I)
    if m:
        school_code = m.group(1).strip()
    if not aff_no:
        lines = [l.strip() for l in aff_text.split("\n") if l.strip()]
        if lines:
            aff_no = lines[0]
        if len(lines) >= 2:
            school_code = lines[1]

    sd_text = txt(2)
    state_val = district_val = ""
    m = re.search(r"State\s*:\s*(.+?)(?:\n|District|$)", sd_text, re.I)
    if m:
        state_val = m.group(1).strip()
    m = re.search(r"District\s*:\s*(.+)", sd_text, re.I)
    if m:
        district_val = m.group(1).strip()

    status = txt(3)

    sh_text = txt(4)
    school_name = principal = ""
    m = re.search(r"Name\s*:\s*(.+?)(?:\n|Head|Principal|$)", sh_text, re.I)
    if m:
        school_name = m.group(1).strip()
    m = re.search(r"(?:Head/?Principal|Principal)\s*Name\s*:?\s*(.+)", sh_text, re.I)
    if m:
        principal = m.group(1).strip()

    addr_text = txt(5)
    address = website = ""
    m = re.search(r"Address\s*:\s*(.+?)(?:\nWebsite|$)", addr_text, re.I | re.S)
    if m:
        address = m.group(1).strip()
    else:
        address = addr_text
    m = re.search(r"Website\s*:\s*(.+)", addr_text, re.I)
    if m:
        website = m.group(1).strip()

    return {
        "S No": s_no,
        "Affiliation No": aff_no,
        "School Code": school_code,
        "State": state_val,
        "District": district_val,
        "Status": status,
        "School Name": school_name,
        "Head/Principal Name": principal,
        "Address": address,
        "Website": website,
    }


def scrape_schools(state, district, progress_callback=None):
    """Full scrape pipeline. Returns list of dicts."""
    driver = _create_headless_driver()
    try:
        if progress_callback:
            progress_callback("Navigating to SARAS website...")

        driver.get(SARAS_URL)
        time.sleep(5)

        # Click "State wise"
        if progress_callback:
            progress_callback("Selecting 'State wise' mode...")
        radio = driver.find_element(By.ID, "SearchMainRadioState_wise")
        _js_click(driver, radio)
        time.sleep(2)

        # Select state
        if progress_callback:
            progress_callback(f"Selecting State: {state}...")
        if not _select_option_by_text(driver, "State", state):
            raise ValueError(f"Could not find State '{state}' in the dropdown.")
        time.sleep(4)

        # Select district
        if progress_callback:
            progress_callback(f"Selecting District: {district}...")
        if not _select_option_by_text(driver, "District", district):
            time.sleep(3)
            if not _select_option_by_text(driver, "District", district):
                raise ValueError(f"Could not find District '{district}' in the dropdown.")
        time.sleep(1)

        # Click SEARCH
        if progress_callback:
            progress_callback("Searching...")
        submitted = False
        for selector in [
            "input[type='submit'][value='SEARCH']",
            "input[type='submit'][value='Search']",
            "input[type='submit']",
            "button[type='submit']",
        ]:
            try:
                btn = driver.find_element(By.CSS_SELECTOR, selector)
                _js_click(driver, btn)
                submitted = True
                break
            except NoSuchElementException:
                continue
        if not submitted:
            for tag in ["button", "input", "a"]:
                for el in driver.find_elements(By.TAG_NAME, tag):
                    txt = (el.text or el.get_attribute("value") or "").strip()
                    if "search" in txt.lower():
                        _js_click(driver, el)
                        submitted = True
                        break
                if submitted:
                    break
        if not submitted:
            driver.execute_script("document.querySelector('form').submit();")

        time.sleep(6)

        # Max entries per page
        try:
            sel = driver.find_element(By.CSS_SELECTOR, "select[name='myTable_length']")
            Select(sel).select_by_value("100")
            time.sleep(3)
        except NoSuchElementException:
            pass

        # Scrape pages
        total = _get_total_entries(driver)
        all_data = []
        page = 1

        while True:
            if progress_callback:
                msg = f"Scraping page {page}..."
                if total:
                    msg += f" ({len(all_data)}/{total} schools)"
                progress_callback(msg)

            rows_data = []
            try:
                tbody = driver.find_element(By.CSS_SELECTOR, "#myTable tbody")
                rows = tbody.find_elements(By.TAG_NAME, "tr")
                for row in rows:
                    if "No data available" in row.text:
                        continue
                    try:
                        parsed = _parse_row(row)
                        if parsed:
                            rows_data.append(parsed)
                    except StaleElementReferenceException:
                        continue
            except NoSuchElementException:
                pass

            if not rows_data:
                break

            all_data.extend(rows_data)

            if total and len(all_data) >= total:
                break

            # Next page
            try:
                next_btn = driver.find_element(By.ID, "myTable_next")
                classes = next_btn.get_attribute("class") or ""
                if "disabled" in classes:
                    break
                _js_click(driver, next_btn)
                time.sleep(2)
                page += 1
            except NoSuchElementException:
                break

        return all_data
    finally:
        driver.quit()


# ──────────────────── Coaching Centres scraper (Google Maps) ───────────────

def _normalize_coaching_name(name):
    """Normalize a coaching centre name for grouping (strip branch/location
    suffixes so 'FIITJEE - Indiranagar' and 'FIITJEE Whitefield' group together)."""
    n = (name or "").strip()
    # Remove anything after a dash, pipe, or parenthesis (usually the branch)
    n = re.split(r"\s[-–|]\s|\s\(|,", n, maxsplit=1)[0]
    # Remove common branch suffix words
    n = re.sub(
        r"\b(branch|centre|center|campus|institute|classes|academy|tuitions?|coaching)\b",
        "",
        n,
        flags=re.I,
    )
    return re.sub(r"\s+", " ", n).strip().upper()


def _scroll_maps_results(driver, max_scrolls=40):
    """Scroll the Google Maps results feed until no more results load."""
    try:
        feed = driver.find_element(By.CSS_SELECTOR, "div[role='feed']")
    except NoSuchElementException:
        return
    last_count = 0
    stagnant = 0
    for _ in range(max_scrolls):
        driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight;", feed)
        time.sleep(2)
        cards = feed.find_elements(By.CSS_SELECTOR, "a.hfpxzc")
        if len(cards) == last_count:
            stagnant += 1
            if stagnant >= 3:
                break
        else:
            stagnant = 0
            last_count = len(cards)
        # Stop if Maps shows the end-of-list sentinel
        if "You've reached the end of the list" in feed.text:
            break


def _parse_maps_card(card):
    """Extract info from a single Maps result card element."""
    try:
        name = card.get_attribute("aria-label") or ""
        name = name.strip()
        link = card.get_attribute("href") or ""
    except StaleElementReferenceException:
        return None

    if not name:
        return None

    parent_text = ""
    try:
        # The enclosing card contains rating, address, phone, type
        parent = card.find_element(By.XPATH, "./ancestor::div[contains(@class,'Nv2PK')][1]")
        parent_text = parent.text
    except NoSuchElementException:
        try:
            parent = card.find_element(By.XPATH, "./..")
            parent_text = parent.text
        except Exception:
            parent_text = ""

    rating = ""
    reviews = ""
    m = re.search(r"(\d\.\d)\s*\(([\d,]+)\)", parent_text)
    if m:
        rating = m.group(1)
        reviews = m.group(2).replace(",", "")
    else:
        m = re.search(r"^(\d\.\d)\s*$", parent_text, re.M)
        if m:
            rating = m.group(1)

    phone = ""
    m = re.search(r"(\+?91[\s\-]?\d{5}[\s\-]?\d{5}|\d{3,5}[\s\-]\d{6,8}|\(\d+\)\s?\d[\d\s\-]{6,})", parent_text)
    if m:
        phone = m.group(1).strip()

    # Address / category line: the text lines between rating and phone
    lines = [l.strip() for l in parent_text.splitlines() if l.strip()]
    address = ""
    category = ""
    for line in lines:
        if line == name:
            continue
        if rating and line.startswith(rating):
            continue
        if phone and phone in line:
            continue
        # A category line usually doesn't contain digits; an address usually does
        if not category and not re.search(r"\d", line) and len(line) < 60:
            category = line
        elif not address and re.search(r"\d", line):
            address = line
        if category and address:
            break

    return {
        "Coaching Centre": name,
        "Category": category,
        "Address": address,
        "Phone": phone,
        "Rating": rating,
        "Reviews": reviews,
        "Maps Link": link,
    }


_COACHING_QUERIES = [
    "coaching classes for class 9 and 10 in {loc}",
    "tuition centres for class 10 in {loc}",
    "CBSE coaching class 9 10 {loc}",
]


def scrape_coaching_centres(state, district, progress_callback=None):
    """Scrape Google Maps for coaching centres teaching Class 9 & 10 in the
    given state/district. Returns a list of dicts with multi-location counts."""
    location = f"{district}, {state}, India"
    driver = _create_headless_driver()
    seen_links = set()
    results = []

    try:
        for i, tpl in enumerate(_COACHING_QUERIES, 1):
            query = tpl.format(loc=location)
            if progress_callback:
                progress_callback(f"Searching Google Maps ({i}/{len(_COACHING_QUERIES)}): '{query}'")

            url = (
                "https://www.google.com/maps/search/"
                + urllib.parse.quote(query)
                + "?hl=en"
            )
            driver.get(url)
            time.sleep(5)

            # Dismiss consent dialog if present
            try:
                btn = driver.find_element(
                    By.XPATH,
                    "//button[.//span[contains(., 'Accept') or contains(., 'I agree')]]",
                )
                _js_click(driver, btn)
                time.sleep(2)
            except NoSuchElementException:
                pass

            if progress_callback:
                progress_callback(f"Loading all results for query {i}...")
            _scroll_maps_results(driver)

            try:
                cards = driver.find_elements(By.CSS_SELECTOR, "a.hfpxzc")
            except Exception:
                cards = []

            for card in cards:
                try:
                    link = card.get_attribute("href") or ""
                except StaleElementReferenceException:
                    continue
                if not link or link in seen_links:
                    continue
                parsed = _parse_maps_card(card)
                if not parsed:
                    continue
                seen_links.add(link)
                results.append(parsed)

            if progress_callback:
                progress_callback(f"Found {len(results)} unique centres so far...")

        # Add branch counts via normalized-name grouping
        name_counts = Counter(_normalize_coaching_name(r["Coaching Centre"]) for r in results)
        for r in results:
            key = _normalize_coaching_name(r["Coaching Centre"])
            count = name_counts[key]
            r["Branches"] = f"{count} location{'s' if count > 1 else ''}"
            r["_group_key"] = key

        # Sort so branches of same chain stay together, chains with more branches first
        results.sort(key=lambda r: (-name_counts[r["_group_key"]], r["_group_key"], r["Address"]))

        # Add serial numbers and strip internal key
        for i, r in enumerate(results, 1):
            r["S No"] = i
            r.pop("_group_key", None)

        return results
    finally:
        driver.quit()


# ──────────────────── Excel generation ─────────────────────────────────────

HEADERS = [
    "S No", "Affiliation No", "School Code", "State", "District",
    "Status", "School Name", "Head/Principal Name", "Address", "Website",
]

COL_WIDTHS = {
    "S No": 8, "Affiliation No": 16, "School Code": 14, "State": 18,
    "District": 22, "Status": 22, "School Name": 42,
    "Head/Principal Name": 28, "Address": 55, "Website": 38,
}


def generate_excel(data):
    """Generate Excel workbook and return as bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "CBSE Schools"

    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    for r, row_data in enumerate(data, 2):
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=r, column=c, value=row_data.get(h, ""))
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border

    for c, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTHS.get(h, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


COACHING_HEADERS = [
    "S No", "Coaching Centre", "Branches", "Category",
    "Address", "Phone", "Rating", "Reviews", "Maps Link",
]

COACHING_COL_WIDTHS = {
    "S No": 6, "Coaching Centre": 36, "Branches": 14, "Category": 22,
    "Address": 55, "Phone": 18, "Rating": 8, "Reviews": 10, "Maps Link": 45,
}


def generate_coaching_excel(data):
    """Generate Excel workbook for coaching centre data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Coaching Centres"

    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_font = Font(name="Calibri", size=10)
    data_align = Alignment(vertical="top", wrap_text=True)

    for col, h in enumerate(COACHING_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    for r, row_data in enumerate(data, 2):
        for c, h in enumerate(COACHING_HEADERS, 1):
            value = row_data.get(h, "")
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border
            if h == "Maps Link" and value:
                cell.hyperlink = value
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    for c, h in enumerate(COACHING_HEADERS, 1):
        ws.column_dimensions[get_column_letter(c)].width = COACHING_COL_WIDTHS.get(h, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ──────────────────── Streamlit UI ─────────────────────────────────────────

# Sidebar
with st.sidebar:
    st.image("https://saras.cbse.gov.in/saras/ui/assets/images/cbse-logo.png", width=80)
    st.markdown("### How to use")
    st.markdown("""
    **CBSE Schools tab**
    1. Select **State** & **District**
    2. Click **🔍 Search Schools**
    3. Download as Excel

    **Coaching Centres tab**
    1. Select **State** & **District**
    2. Click **🔍 Search Coaching**
    3. Click any Maps link for full details
    """)
    st.divider()
    st.markdown(
        "**Schools source:** [SARAS 7.0](https://saras.cbse.gov.in/saras/AffiliatedList/ListOfSchdirReport)"
    )
    st.markdown("**Coaching source:** Google Maps")
    st.markdown(f"**Total CBSE schools:** 32,892")

# Load state/district data
with st.spinner("Loading states and districts from CBSE..."):
    try:
        state_map, district_map = fetch_states_and_districts()
        states_loaded = True
    except Exception as e:
        st.error(f"Failed to load state data: {e}")
        states_loaded = False

if states_loaded:
    tab_schools, tab_coaching = st.tabs(["🏫 CBSE Schools", "📚 Coaching Centres"])

    # ══════════════════════ TAB 1: SCHOOLS ═════════════════════════════════
    with tab_schools:
        col1, col2, col3 = st.columns([2, 2, 1])

        with col1:
            selected_state = st.selectbox(
                "🗺️ Select State",
                options=[""] + sorted(state_map.keys()),
                index=0,
                placeholder="Choose a state...",
                key="school_state",
            )

        with col2:
            if selected_state and selected_state in district_map:
                districts = district_map[selected_state]
            else:
                districts = []
            selected_district = st.selectbox(
                "📍 Select District",
                options=[""] + districts,
                index=0,
                placeholder="Choose a district...",
                disabled=not selected_state,
                key="school_district",
            )

        with col3:
            st.markdown("<br>", unsafe_allow_html=True)
            search_clicked = st.button(
                "🔍 Search Schools",
                type="primary",
                use_container_width=True,
                disabled=not (selected_state and selected_district),
                key="school_btn",
            )

        if search_clicked and selected_state and selected_district:
            status_container = st.empty()
            progress_bar = st.progress(0)

            step_count = [0]

            def update_status(msg):
                step_count[0] += 1
                progress_val = min(step_count[0] * 12, 95)
                progress_bar.progress(progress_val)
                status_container.info(f"⏳ {msg}")

            try:
                data = scrape_schools(selected_state, selected_district, update_status)
                progress_bar.progress(100)

                if data:
                    status_container.success(
                        f"✅ Found **{len(data)}** schools in **{selected_district}**, **{selected_state}**"
                    )

                    st.divider()
                    c1, c2, c3, c4 = st.columns(4)
                    with c1:
                        st.markdown(
                            f"<div class='stat-card'><h3>Total Schools</h3><p>{len(data)}</p></div>",
                            unsafe_allow_html=True,
                        )
                    with c2:
                        secondary = sum(1 for d in data if "secondary" in d.get("Status", "").lower() and "senior" not in d.get("Status", "").lower())
                        st.markdown(
                            f"<div class='stat-card'><h3>Secondary</h3><p>{secondary}</p></div>",
                            unsafe_allow_html=True,
                        )
                    with c3:
                        senior = sum(1 for d in data if "senior" in d.get("Status", "").lower())
                        st.markdown(
                            f"<div class='stat-card'><h3>Sr. Secondary</h3><p>{senior}</p></div>",
                            unsafe_allow_html=True,
                        )
                    with c4:
                        with_website = sum(1 for d in data if d.get("Website"))
                        st.markdown(
                            f"<div class='stat-card'><h3>With Website</h3><p>{with_website}</p></div>",
                            unsafe_allow_html=True,
                        )

                    st.divider()
                    st.markdown("### 📋 School Data")

                    search_term = st.text_input(
                        "🔎 Filter results",
                        placeholder="Type school name, principal, address...",
                        key="school_filter",
                    )

                    display_data = data
                    if search_term:
                        term = search_term.upper()
                        display_data = [
                            d for d in data
                            if any(term in str(v).upper() for v in d.values())
                        ]
                        st.caption(f"Showing {len(display_data)} of {len(data)} schools")

                    st.dataframe(
                        display_data,
                        use_container_width=True,
                        height=400,
                        column_config={
                            "S No": st.column_config.NumberColumn("S No", width="small"),
                            "Website": st.column_config.LinkColumn("Website"),
                        },
                    )

                    excel_bytes = generate_excel(data)
                    filename = f"CBSE_Schools_{selected_state.replace(' ', '_')}_{selected_district.replace(' ', '_')}.xlsx"

                    st.download_button(
                        label="📥 Download Excel File",
                        data=excel_bytes,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                        key="school_dl",
                    )
                else:
                    status_container.warning("No schools found for the selected State and District.")
                    progress_bar.empty()

            except ValueError as e:
                status_container.error(f"❌ {e}")
                progress_bar.empty()
            except Exception as e:
                status_container.error(f"❌ An error occurred: {e}")
                progress_bar.empty()
        else:
            st.info("👆 Select a State and District, then click **Search Schools** to begin.")

    # ══════════════════════ TAB 2: COACHING CENTRES ════════════════════════
    with tab_coaching:
        st.markdown(
            "Find coaching centres teaching **Class 9 & Class 10** students in your area. "
            "Results include multi-location chains and direct Google Maps links for contact details."
        )

        cc1, cc2, cc3 = st.columns([2, 2, 1])

        with cc1:
            coaching_state = st.selectbox(
                "🗺️ Select State",
                options=[""] + sorted(state_map.keys()),
                index=0,
                placeholder="Choose a state...",
                key="coaching_state",
            )

        with cc2:
            if coaching_state and coaching_state in district_map:
                coaching_districts = district_map[coaching_state]
            else:
                coaching_districts = []
            coaching_district = st.selectbox(
                "📍 Select District",
                options=[""] + coaching_districts,
                index=0,
                placeholder="Choose a district...",
                disabled=not coaching_state,
                key="coaching_district",
            )

        with cc3:
            st.markdown("<br>", unsafe_allow_html=True)
            coaching_clicked = st.button(
                "🔍 Search Coaching",
                type="primary",
                use_container_width=True,
                disabled=not (coaching_state and coaching_district),
                key="coaching_btn",
            )

        if coaching_clicked and coaching_state and coaching_district:
            c_status = st.empty()
            c_progress = st.progress(0)
            c_step = [0]

            def coaching_status(msg):
                c_step[0] += 1
                c_progress.progress(min(c_step[0] * 8, 95))
                c_status.info(f"⏳ {msg}")

            try:
                coaching_data = scrape_coaching_centres(
                    coaching_state, coaching_district, coaching_status
                )
                c_progress.progress(100)

                if coaching_data:
                    # Count unique chains (groups) vs total branches
                    unique_chains = len({
                        _normalize_coaching_name(d["Coaching Centre"]) for d in coaching_data
                    })
                    multi_branch = sum(
                        1 for d in coaching_data
                        if d.get("Branches", "").split()[0].isdigit()
                        and int(d["Branches"].split()[0]) > 1
                    )
                    with_phone = sum(1 for d in coaching_data if d.get("Phone"))

                    c_status.success(
                        f"✅ Found **{len(coaching_data)}** coaching locations "
                        f"(**{unique_chains}** unique centres) in "
                        f"**{coaching_district}**, **{coaching_state}**"
                    )

                    st.divider()
                    s1, s2, s3, s4 = st.columns(4)
                    with s1:
                        st.markdown(
                            f"<div class='stat-card'><h3>Total Locations</h3><p>{len(coaching_data)}</p></div>",
                            unsafe_allow_html=True,
                        )
                    with s2:
                        st.markdown(
                            f"<div class='stat-card'><h3>Unique Centres</h3><p>{unique_chains}</p></div>",
                            unsafe_allow_html=True,
                        )
                    with s3:
                        st.markdown(
                            f"<div class='stat-card'><h3>Multi-Branch Rows</h3><p>{multi_branch}</p></div>",
                            unsafe_allow_html=True,
                        )
                    with s4:
                        st.markdown(
                            f"<div class='stat-card'><h3>With Phone</h3><p>{with_phone}</p></div>",
                            unsafe_allow_html=True,
                        )

                    st.divider()
                    st.markdown("### 📋 Coaching Centres")
                    st.caption(
                        "💡 Branches of the same coaching chain are grouped together. "
                        "Click a **Maps Link** to open the listing for phone, website & directions."
                    )

                    c_filter = st.text_input(
                        "🔎 Filter results",
                        placeholder="Type coaching name, area, category...",
                        key="coaching_filter",
                    )

                    display_coaching = coaching_data
                    if c_filter:
                        term = c_filter.upper()
                        display_coaching = [
                            d for d in coaching_data
                            if any(term in str(v).upper() for v in d.values())
                        ]
                        st.caption(f"Showing {len(display_coaching)} of {len(coaching_data)} centres")

                    st.dataframe(
                        display_coaching,
                        use_container_width=True,
                        height=450,
                        column_order=[
                            "S No", "Coaching Centre", "Branches", "Category",
                            "Address", "Phone","Maps Link", "Rating", "Reviews", 
                        ],
                        column_config={
                            "S No": st.column_config.NumberColumn("S No", width="small"),
                            "Coaching Centre": st.column_config.TextColumn("Coaching Centre", width="medium"),
                            "Branches": st.column_config.TextColumn("Branches", width="small"),
                            "Rating": st.column_config.TextColumn("⭐ Rating", width="small"),
                            "Maps Link": st.column_config.LinkColumn(
                                "Maps Link", display_text="🗺️ Open in Maps"
                            ),
                        },
                    )

                    coaching_xlsx = generate_coaching_excel(coaching_data)
                    c_filename = (
                        f"Coaching_Centres_{coaching_state.replace(' ', '_')}_"
                        f"{coaching_district.replace(' ', '_')}.xlsx"
                    )

                    st.download_button(
                        label="📥 Download Excel File",
                        data=coaching_xlsx,
                        file_name=c_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                        key="coaching_dl",
                    )
                else:
                    c_status.warning("No coaching centres found for the selected area.")
                    c_progress.empty()

            except Exception as e:
                c_status.error(f"❌ An error occurred: {e}")
                c_progress.empty()
        else:
            st.info("👆 Select a State and District, then click **Search Coaching** to begin.")
